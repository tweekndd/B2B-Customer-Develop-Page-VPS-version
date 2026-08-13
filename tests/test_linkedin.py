"""
LinkedIn 公司主页发现测试（V5.1）
覆盖：URL 标准化/过滤、vanity 提取、候选评分、搜索发现（mock）、
手动新增、确认互斥、删除、Excel 导出回填
"""
import datetime
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session

from app.database import Base, get_db, Customer, CustomerSocialProfile
from app.auth import hash_password
from main import app


_TEST_DB = os.path.join(os.path.dirname(__file__), "test_api.db")
TEST_DATABASE_URL = f"sqlite:///{_TEST_DB}"
_test_engine = create_engine(TEST_DATABASE_URL, connect_args={"check_same_thread": False})
TestSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=_test_engine)


def override_get_db():
    db = TestSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)


@pytest.fixture(autouse=True)
def setup_db():
    Base.metadata.create_all(bind=_test_engine)
    yield
    Base.metadata.drop_all(bind=_test_engine)


@pytest.fixture(autouse=True)
def login_user(setup_db):
    db = TestSessionLocal()
    try:
        from app.database import User
        db.add(User(
            username="testadmin",
            password_hash=hash_password("testpassword"),
            role="admin",
            is_active=1,
        ))
        db.commit()
    finally:
        db.close()
    resp = client.post(
        "/api/auth/login",
        json={"username": "testadmin", "password": "testpassword"},
    )
    assert resp.status_code == 200, resp.text
    yield
    client.post("/api/auth/logout")


def _create_customer(db: Session, **overrides) -> Customer:
    data = {
        "company_name": "AquaTech Solutions",
        "website": "aquatech-solutions.com",
        "country": "Saudi Arabia",
        "discovery_keyword": "water treatment",
    }
    data.update(overrides)
    c = Customer(**data)
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


class TestUrlNormalization:
    def test_normalize_company_url(self):
        """标准化：去协议差异、追踪参数、尾斜杠"""
        from app.services.linkedin_service import normalize_company_url
        assert normalize_company_url("https://www.linkedin.com/company/aquatech") == \
            "https://www.linkedin.com/company/aquatech"
        assert normalize_company_url("https://linkedin.com/company/aquatech?trk=foo") == \
            "https://www.linkedin.com/company/aquatech"
        assert normalize_company_url("https://www.linkedin.com/company/aquatech/") == \
            "https://www.linkedin.com/company/aquatech"
        assert normalize_company_url("www.linkedin.com/company/aquatech") == \
            "https://www.linkedin.com/company/aquatech"

    def test_reject_personal_pages(self):
        """个人页/职位页/帖子页被过滤"""
        from app.services.linkedin_service import normalize_company_url
        assert normalize_company_url("https://www.linkedin.com/in/johnsmith") is None
        assert normalize_company_url("https://www.linkedin.com/jobs/view/123") is None
        assert normalize_company_url("https://www.linkedin.com/feed/") is None
        assert normalize_company_url("https://www.linkedin.com/learning/courses") is None
        assert normalize_company_url("https://example.com/company/x") is None
        assert normalize_company_url("") is None
        assert normalize_company_url("not-a-url") is None

    def test_extract_vanity_name(self):
        from app.services.linkedin_service import extract_vanity_name
        assert extract_vanity_name("https://www.linkedin.com/company/aquatech/") == "aquatech"
        assert extract_vanity_name("https://www.linkedin.com/company/AquaTech-Co") == "AquaTech-Co"
        assert extract_vanity_name("https://www.linkedin.com/in/john") is None


class TestCandidateScoring:
    def test_domain_match_scores_high(self):
        """域名根命中 +50"""
        from app.services.linkedin_service import score_company_page_candidate
        c = _create_customer(TestSessionLocal())
        cand = {
            "profile_url": "https://www.linkedin.com/company/aquatech-solutions",
            "title": "AquaTech Solutions | LinkedIn",
            "snippet": "",
        }
        score = score_company_page_candidate(cand, c)
        assert score >= 50
        assert score <= 100

    def test_company_name_and_country_hits(self):
        """公司名 +30、国家 +10、关键词 +10"""
        from app.services.linkedin_service import score_company_page_candidate
        c = _create_customer(TestSessionLocal(), website="unknown-domain.example")
        cand = {
            "profile_url": "https://www.linkedin.com/company/whatever",
            "title": "AquaTech Solutions - Saudi Arabia",
            "snippet": "water treatment contractor in Riyadh",
        }
        score = score_company_page_candidate(cand, c)
        # 域名根不匹配（unknown-domain 不在），其余三项应命中
        assert score >= 50  # 30 + 10 + 10

    def test_no_match_scores_zero(self):
        """完全不相关候选得分低"""
        from app.services.linkedin_service import score_company_page_candidate
        c = _create_customer(TestSessionLocal(), website="unrelated.example")
        cand = {
            "profile_url": "https://www.linkedin.com/company/xyz-corp",
            "title": "XYZ Corp",
            "snippet": "finance consulting",
        }
        score = score_company_page_candidate(cand, c)
        assert score < 50


class TestDiscoveryService:
    def test_discover_returns_only_company_pages(self, monkeypatch):
        """搜索发现：仅保留公司页，过滤个人页，且按置信度排序"""
        import asyncio
        import app.services.google_discovery as gd

        async def fake_search(keyword, country, max_results=20, user_id=None, db=None):
            return [
                {"title": "AquaTech Solutions | LinkedIn", "website": "https://www.linkedin.com/company/aquatech-solutions", "snippet": "water treatment"},
                {"title": "John Smith | LinkedIn", "website": "https://www.linkedin.com/in/johnsmith", "snippet": ""},
                {"title": "AquaTech Solutions Careers", "website": "https://www.linkedin.com/company/aquatech-solutions/jobs", "snippet": ""},
                {"title": "Something Else", "website": "https://example.com/not-linkedin", "snippet": ""},
                {"title": "AquaTech | LinkedIn", "website": "https://www.linkedin.com/company/aquatech-solutions?trk=param", "snippet": "water treatment company in Saudi Arabia"},
            ]
        monkeypatch.setattr(gd, "search_google", fake_search)

        from app.services.linkedin_service import discover_company_pages
        results = asyncio.run(discover_company_pages(
            company_name="AquaTech Solutions",
            website="aquatech-solutions.com",
            country="Saudi Arabia",
        ))
        assert len(results) == 1  # 去重后只有公司页
        assert results[0]["profile_url"] == "https://www.linkedin.com/company/aquatech-solutions"
        assert results[0]["vanity_name"] == "aquatech-solutions"
        assert results[0]["confidence"] >= 50

    def test_discover_no_engine_returns_empty(self, monkeypatch):
        """无搜索引擎时返回空列表"""
        import asyncio
        import app.services.google_discovery as gd

        async def fake_search(keyword, country, max_results=20, user_id=None, db=None):
            return []
        monkeypatch.setattr(gd, "search_google", fake_search)

        from app.services.linkedin_service import discover_company_pages
        results = asyncio.run(discover_company_pages(company_name="Foo", website="foo.com"))
        assert results == []


class TestSocialProfileAPI:
    def test_manual_add(self):
        """手动新增 LinkedIn 主页"""
        db = TestSessionLocal()
        c = _create_customer(db)
        db.close()

        resp = client.post(
            f"/api/customers/{c.id}/social-profiles",
            json={"profile_url": "https://www.linkedin.com/company/aquatech-solutions"},
        )
        assert resp.status_code == 200, resp.text
        data = resp.json()
        assert data["profile"]["profile_url"] == "https://www.linkedin.com/company/aquatech-solutions"
        assert data["profile"]["source"] == "manual"
        assert data["profile"]["is_verified"] is False

    def test_manual_add_invalid_url(self):
        """非公司页 URL 返回 400"""
        db = TestSessionLocal()
        c = _create_customer(db)
        db.close()
        resp = client.post(
            f"/api/customers/{c.id}/social-profiles",
            json={"profile_url": "https://www.linkedin.com/in/john"},
        )
        assert resp.status_code == 400

    def test_manual_add_idempotent(self):
        """重复添加同一主页不产生重复记录"""
        db = TestSessionLocal()
        c = _create_customer(db)
        db.close()
        url = "https://www.linkedin.com/company/aquatech-solutions"
        client.post(f"/api/customers/{c.id}/social-profiles", json={"profile_url": url})
        client.post(f"/api/customers/{c.id}/social-profiles", json={"profile_url": url + "?trk=dup"})

        db = TestSessionLocal()
        count = db.query(CustomerSocialProfile).filter(CustomerSocialProfile.customer_id == c.id).count()
        assert count == 1
        db.close()

    def test_verify_single(self):
        """确认主页：同客户只有一个已确认"""
        db = TestSessionLocal()
        c = _create_customer(db)
        db.close()
        r1 = client.post(f"/api/customers/{c.id}/social-profiles",
                         json={"profile_url": "https://www.linkedin.com/company/a"})
        r2 = client.post(f"/api/customers/{c.id}/social-profiles",
                         json={"profile_url": "https://www.linkedin.com/company/b"})
        id_a = r1.json()["profile"]["id"]
        id_b = r2.json()["profile"]["id"]

        resp = client.put(f"/api/social-profiles/{id_a}", json={"is_verified": True})
        assert resp.status_code == 200
        resp = client.put(f"/api/social-profiles/{id_b}", json={"is_verified": True})
        assert resp.status_code == 200

        db = TestSessionLocal()
        profiles = db.query(CustomerSocialProfile).filter(CustomerSocialProfile.customer_id == c.id).all()
        verified = [p for p in profiles if p.is_verified]
        assert len(verified) == 1
        assert verified[0].id == id_b  # 后确认者胜出
        db.close()

    def test_delete(self):
        """删除候选主页"""
        db = TestSessionLocal()
        c = _create_customer(db)
        db.close()
        r = client.post(f"/api/customers/{c.id}/social-profiles",
                        json={"profile_url": "https://www.linkedin.com/company/a"})
        pid = r.json()["profile"]["id"]

        resp = client.delete(f"/api/social-profiles/{pid}")
        assert resp.status_code == 200

        db = TestSessionLocal()
        assert db.query(CustomerSocialProfile).filter(CustomerSocialProfile.id == pid).count() == 0
        db.close()

    def test_list_verified_first(self):
        """列表：已确认优先展示"""
        db = TestSessionLocal()
        c = _create_customer(db)
        db.close()
        r1 = client.post(f"/api/customers/{c.id}/social-profiles",
                         json={"profile_url": "https://www.linkedin.com/company/a"})
        r2 = client.post(f"/api/customers/{c.id}/social-profiles",
                         json={"profile_url": "https://www.linkedin.com/company/b"})
        client.put(f"/api/social-profiles/{r1.json()['profile']['id']}", json={"is_verified": True})

        resp = client.get(f"/api/customers/{c.id}/social-profiles")
        assert resp.status_code == 200
        profiles = resp.json()["profiles"]
        assert len(profiles) == 2
        assert profiles[0]["is_verified"] is True
        assert profiles[0]["id"] == r1.json()["profile"]["id"]


class TestExcelExportLinkedIn:
    def test_export_verified_linkedin_column(self):
        """Excel 导出 H 列回填已确认 LinkedIn"""
        import io
        import openpyxl
        db = TestSessionLocal()
        c = _create_customer(db)
        db.close()
        client.post(f"/api/customers/{c.id}/social-profiles",
                    json={"profile_url": "https://www.linkedin.com/company/aquatech-solutions"})
        profiles = client.get(f"/api/customers/{c.id}/social-profiles").json()["profiles"]
        client.put(f"/api/social-profiles/{profiles[0]['id']}", json={"is_verified": True})

        resp = client.get("/api/export-excel")
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        assert "领英" in header
        h_idx = header.index("领英") + 1
        values = [ws.cell(row=r, column=h_idx).value for r in range(2, ws.max_row + 1)]
        assert values and values[0] == "https://www.linkedin.com/company/aquatech-solutions"
